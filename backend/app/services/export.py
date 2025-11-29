"""
Export services for Excel and PDF reports.
"""
import os
from pathlib import Path
from typing import List
from sqlalchemy.orm import Session
from uuid import UUID
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import time
import logging
from app.models import AuditRun, LaneStat, AuditResult, Shipment
from app.services.audit_engine import get_exceptions
from app.services.llm_reports import generate_executive_summary

# Use absolute path for exports directory
EXPORT_DIR = Path(__file__).parent.parent.parent / "exports"
EXPORT_DIR.mkdir(exist_ok=True)
logger = logging.getLogger(__name__)


def generate_excel_report(db: Session, audit_run_id: UUID) -> str:
    """
    Generate Excel report with:
    - Summary sheet
    - Lane statistics
    - Exception shipments
    """
    start_time = time.perf_counter()
    audit_run = db.query(AuditRun).filter(AuditRun.id == audit_run_id).first()
    if not audit_run:
        raise ValueError(f"Audit run {audit_run_id} not found")
    
    wb = Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    summary_metrics = audit_run.summary_metrics or {}
    ws_summary.append(["Audit Summary"])
    ws_summary.append(["Customer", audit_run.customer.name])
    ws_summary.append(["Audit Name", audit_run.name])
    ws_summary.append(["Period", audit_run.label or ""])
    ws_summary.append([])
    ws_summary.append(["Key Metrics"])
    ws_summary.append(["Total Shipments", summary_metrics.get("shipment_count", 0)])
    ws_summary.append(["Total Spend", f"${summary_metrics.get('total_spend', 0):,.2f}"])
    ws_summary.append(["Average Cost per Shipment", f"${summary_metrics.get('avg_cost_per_shipment', 0):,.2f}"])
    ws_summary.append(["Average Cost per LB", f"${summary_metrics.get('avg_cost_per_lb', 0):,.4f}"])
    ws_summary.append(["Average Cost per Pallet", f"${summary_metrics.get('avg_cost_per_pallet', 0):,.2f}"])
    
    # Lane statistics sheet
    ws_lanes = wb.create_sheet("Lane Statistics")
    lane_stats = db.query(LaneStat).filter(
        LaneStat.audit_run_id == audit_run_id
    ).order_by(LaneStat.total_spend.desc()).all()
    
    ws_lanes.append([
        "Origin DC", "Dest Province", "Dest Region", "Dest City",
        "Shipments", "Total Spend", "Total Weight", "Total Pallets",
        "Avg $/LB", "Avg $/Pallet", "Theoretical Best", "Savings", "Savings %"
    ])
    
    for ls in lane_stats:
        ws_lanes.append([
            ls.origin_dc or "",
            ls.dest_province or "",
            ls.dest_region or "",
            ls.dest_city or "",
            ls.shipment_count,
            float(ls.total_spend),
            float(ls.total_weight),
            float(ls.total_pallets),
            float(ls.avg_cost_per_lb) if ls.avg_cost_per_lb else "",
            float(ls.avg_cost_per_pallet) if ls.avg_cost_per_pallet else "",
            float(ls.theoretical_best_spend) if ls.theoretical_best_spend else "",
            float(ls.theoretical_savings) if ls.theoretical_savings else "",
            f"{float(ls.savings_pct):.2f}%" if ls.savings_pct else "",
        ])
    
    # Exceptions sheet
    ws_exceptions = wb.create_sheet("Exceptions")
    exceptions = get_exceptions(db, audit_run_id, "all")
    
    ws_exceptions.append([
        "Shipment Ref", "Origin DC", "Dest City", "Dest Province",
        "Weight", "Pallets", "Charge", "Cost/LB", "Flags"
    ])
    
    for exc in exceptions:
        ws_exceptions.append([
            exc.get("shipment_ref", ""),
            exc.get("origin_dc", ""),
            exc.get("dest_city", ""),
            exc.get("dest_province", ""),
            exc.get("weight"),
            exc.get("pallets"),
            exc.get("actual_charge"),
            exc.get("cost_per_lb"),
            ", ".join(exc.get("flags", [])),
        ])
    
    # Save file
    file_path = EXPORT_DIR / f"audit_report_{audit_run_id}.xlsx"
    wb.save(file_path)
    duration = round(time.perf_counter() - start_time, 3)
    logger.info(
        "Excel report generated for audit %s lanes=%d exceptions=%d in %.2fs",
        audit_run_id,
        len(lane_stats),
        len(exceptions),
        duration,
    )
    
    return str(file_path)


def generate_pdf_report(db: Session, audit_run_id: UUID) -> str:
    """
    Generate PDF executive summary.
    """
    start_time = time.perf_counter()
    audit_run = db.query(AuditRun).filter(AuditRun.id == audit_run_id).first()
    if not audit_run:
        raise ValueError(f"Audit run {audit_run_id} not found")
    
    # Generate summary text
    summary_text = generate_executive_summary(db, audit_run_id)
    
    # Create PDF
    file_path = EXPORT_DIR / f"audit_summary_{audit_run_id}.pdf"
    doc = SimpleDocTemplate(str(file_path), pagesize=letter)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
    )
    
    # Title
    story.append(Paragraph(f"Freight Audit Report: {audit_run.name}", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Summary text (split by lines and convert to paragraphs)
    for line in summary_text.split('\n'):
        if line.strip():
            if line.startswith('#'):
                # Heading
                level = len(line) - len(line.lstrip('#'))
                style = styles[f'Heading{min(level, 6)}']
                text = line.lstrip('#').strip()
                story.append(Paragraph(text, style))
            else:
                story.append(Paragraph(line, styles['Normal']))
        else:
            story.append(Spacer(1, 0.1*inch))
    
    doc.build(story)
    duration = round(time.perf_counter() - start_time, 3)
    logger.info("PDF summary generated for audit %s in %.2fs", audit_run_id, duration)
    
    return str(file_path)

